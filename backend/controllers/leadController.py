import csv
import io
import re
import traceback
import os
from flask import request, jsonify, g
from werkzeug.utils import secure_filename
from models.leadModel import (
    LEAD_FIELDS,
    create_batch,
    existing_email_keys,
    get_batch,
    list_leads,
    mark_leads_qualified,
    sanitize_batch_doc,
    sanitize_lead_doc,
    save_leads_for_batch,
    update_batch_status,
    get_latest_batch_leads,
    get_latest_batch,
)
from models.qualificationResultModel import get_qualification_results, sanitize_qualification_result
from Services.qualificationService import LeadQualificationService


HEADER_ALIASES = {
    "name": "Name",
    "full name": "Name",
    "contact name": "Name",
    "company": "Company",
    "company name": "Company",
    "account": "Company",
    "role": "Role",
    "title": "Role",
    "job title": "Role",
    "industry": "Industry",
    "email": "Email",
    "email address": "Email",
    "work email": "Email",
    "employee count": "Employee Count",
    "employees": "Employee Count",
    "company size": "Employee Count",
    "headcount": "Employee Count",
    "company tech stack": "Company Tech Stack",
    "tech stack": "Company Tech Stack",
    "technologies": "Company Tech Stack",
    "technology": "Company Tech Stack",
    "tools": "Company Tech Stack",
    "stack": "Company Tech Stack",
    "funding stage": "Funding Stage",
    "funding": "Funding Stage",
    "stage": "Funding Stage",
    "location": "Location",
    "recent trigger event": "Recent Trigger Event",
    "recent trigger": "Recent Trigger Event",
    "trigger event": "Recent Trigger Event",
    "target role match": "Target Role Match",
    "role match": "Target Role Match",
    "target role": "Target Role Match",
    "currently hiring": "Currently Hiring",
    "hiring": "Currently Hiring",
}


EMAIL_RE = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")
CSV_EXTENSIONS = {".csv"}
EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}

ADDITIONAL_LEAD_FIELDS = [
    "Funding Stage",
    "Location",
    "Recent Trigger Event",
    "Target Role Match",
]

PERMITTED_LEAD_FIELDS = LEAD_FIELDS + ADDITIONAL_LEAD_FIELDS


def _clean(value):
    return " ".join(str(value or "").strip().split())


def _normalize_header(value):
    key = _clean(value).lower().replace("_", " ").replace("-", " ")
    return HEADER_ALIASES.get(" ".join(key.split()))


def _normalize_employee_count(value):
    cleaned = _clean(value).replace(",", "")
    match = re.search(r"\d+", cleaned)
    return match.group(0) if match else ""


def _normalize_currently_hiring(value):
    normalized = _clean(value).lower()
    if normalized in {"yes", "y", "true", "1", "hiring"}:
        return "Yes"
    if normalized in {"no", "n", "false", "0", "not hiring"}:
        return "No"
    return ""


def _rows_to_leads(fieldnames, rows):
    header_map = {}
    for header in fieldnames:
        normalized = _normalize_header(header)
        if normalized:
            header_map[normalized] = header

    missing = [field for field in LEAD_FIELDS if field not in header_map]
    if missing:
        return [], [f"Missing columns: {', '.join(missing)}"], 0

    leads = []
    errors = []
    seen_emails = set()
    duplicate_rows = 0

    for index, row in enumerate(rows, start=2):
        lead = {
            "name": _clean(row.get(header_map["Name"])),
            "company": _clean(row.get(header_map["Company"])),
            "role": _clean(row.get(header_map["Role"])),
            "industry": _clean(row.get(header_map["Industry"])),
            "email": _clean(row.get(header_map["Email"])).lower(),
            "employeeCount": _normalize_employee_count(row.get(header_map["Employee Count"])),
            "companyTechStack": _clean(row.get(header_map["Company Tech Stack"])),
            "fundingStage": _clean(row.get(header_map["Funding Stage"])) if "Funding Stage" in header_map else "",
            "location": _clean(row.get(header_map["Location"])) if "Location" in header_map else "",
            "recentTriggerEvent": _clean(row.get(header_map["Recent Trigger Event"])) if "Recent Trigger Event" in header_map else "",
            "targetRoleMatch": _clean(row.get(header_map["Target Role Match"])) if "Target Role Match" in header_map else "",
            "currentlyHiring": _normalize_currently_hiring(row.get(header_map["Currently Hiring"])) if "Currently Hiring" in header_map else "",
            "duplicateStatus": "new",
        }

        if not any(lead.values()):
            continue

        row_errors = []
        if not lead["name"]:
            row_errors.append("name is required")
        if not lead["company"]:
            row_errors.append("company is required")
        if not EMAIL_RE.match(lead["email"]):
            row_errors.append("email is invalid")
        if not lead["employeeCount"]:
            row_errors.append("employee count is required")
        if not lead["companyTechStack"]:
            row_errors.append("company tech stack is required")
        if not lead["currentlyHiring"]:
            row_errors.append("currently hiring is required")

        if row_errors:
            errors.append(f"Row {index}: {', '.join(row_errors)}.")
            continue

        email_key = lead["email"]
        if email_key in seen_emails:
            duplicate_rows += 1
            continue

        seen_emails.add(email_key)
        leads.append(lead)

    if not leads and not errors:
        errors.append("CSV must include at least one lead.")

    return leads, errors, duplicate_rows


def _parse_csv(file_storage):
    raw = file_storage.read()
    text = raw.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))

    if not reader.fieldnames:
        return [], ["CSV must include a header row."], 0

    return _rows_to_leads(reader.fieldnames, reader)


def _parse_excel(file_storage):
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], ["Excel uploads require the openpyxl package to be installed."], 0

    file_storage.stream.seek(0)
    workbook = load_workbook(file_storage.stream, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], ["Excel file must include a header row."], 0

    headers = [_clean(value) for value in rows[0]]
    if not any(headers):
        return [], ["Excel file must include a header row."], 0

    dict_rows = []
    for values in rows[1:]:
        row = {}
        for index, header in enumerate(headers):
            if header:
                row[header] = values[index] if index < len(values) else ""
        dict_rows.append(row)

    return _rows_to_leads(headers, dict_rows)


def _file_extension(file_name):
    lowered = file_name.lower()
    if "." not in lowered:
        return ""
    return lowered[lowered.rfind("."):]


def _parse_upload(file_storage, file_name):
    extension = _file_extension(file_name)
    if extension in CSV_EXTENSIONS:
        return _parse_csv(file_storage)
    if extension in EXCEL_EXTENSIONS:
        return _parse_excel(file_storage)
    return [], ["Only CSV and Excel .xlsx/.xlsm files are supported."], 0


def upload_leads():
    try:
        file_storage = request.files.get("file")
        if not file_storage:
            return jsonify({"error": "Lead file is required"}), 400

        file_name = secure_filename(file_storage.filename or "uploaded-leads.csv")

        leads, errors, duplicate_rows = _parse_upload(file_storage, file_name)

        if errors:
            return jsonify({"errors": errors}), 400

        existing_emails = existing_email_keys(g.current_user["id"], [lead["email"] for lead in leads])
        for lead in leads:
            if lead["email"] in existing_emails:
                lead["duplicateStatus"] = "updated"

        stats = {
            "uploadedRows": len(leads) + duplicate_rows,
            "savedRows": len(leads),
            "duplicateRows": duplicate_rows,
            "updatedRows": len(existing_emails),
        }
        batch = create_batch(g.current_user["id"], file_name, stats)
        saved_leads = save_leads_for_batch(g.current_user["id"], batch["_id"], leads)

        return jsonify(
            {
                "batch": sanitize_batch_doc(batch),
                "leads": [sanitize_lead_doc(lead) for lead in saved_leads],
            }
        ), 201

    except UnicodeDecodeError:
        return jsonify({"error": "CSV must be UTF-8 encoded"}), 400
    except Exception:
        traceback.print_exc()
        return jsonify({"error": "Internal server error"}), 500


def get_leads():
    try:
        batch_id = request.args.get("batchId")
        lead_id = request.args.get("leadId") or request.args.get("id")
        leads = list_leads(g.current_user["id"], batch_id=batch_id, lead_id=lead_id)
        return jsonify({"leads": [sanitize_lead_doc(lead) for lead in leads]}), 200
    except Exception:
        traceback.print_exc()
        return jsonify({"error": "Internal server error"}), 500


def _employee_size_score(employee_count):
    count = int(employee_count or 0)
    if 51 <= count <= 500:
        return 25, "Company size sits in the strongest mid-market buying window."
    if 11 <= count <= 1000:
        return 15, "Company size is workable for outbound qualification."
    return 6, "Company size is outside the preferred ICP center."


def _agent_research(lead):
    score = 35
    reasons = []

    role = (lead.get("Role") or "").lower()
    industry = (lead.get("Industry") or "").lower()
    tech_stack = (lead.get("Company Tech Stack") or "").lower()
    company = lead.get("Company") or "the account"

    decision_terms = ["founder", "ceo", "vp", "head", "director", "owner", "chief"]
    if any(term in role for term in decision_terms):
        score += 25
        reasons.append("Contact role has clear decision-maker or budget-owner signal.")
    else:
        reasons.append("Contact may need additional stakeholder mapping.")

    priority_industries = ["saas", "ai", "software", "fintech", "cybersecurity", "e-commerce"]
    if any(term in industry for term in priority_industries):
        score += 20
        reasons.append("Industry matches a high-priority digital buying segment.")
    else:
        score += 8
        reasons.append("Industry is valid but not a top-priority segment.")

    size_score, size_reason = _employee_size_score(lead.get("Employee Count"))
    score += size_score
    reasons.append(size_reason)

    priority_stack = ["salesforce", "hubspot", "marketo", "apollo", "outreach", "salesloft", "segment", "snowflake", "aws", "openai"]
    if any(tool in tech_stack for tool in priority_stack):
        score += 12
        reasons.append("Company tech stack shows strong compatibility with AI-led sales workflows.")
    else:
        score += 4
        reasons.append("Tech stack is captured but has limited high-intent sales tooling signal.")

    score = min(score, 100)
    qualified = score >= 70
    research_summary = (
        f"Agent reviewed {company}, role seniority, industry fit, employee count, and company tech stack. "
        f"The lead scored {score}/100 and is {'qualified' if qualified else 'not qualified yet'}."
    )

    return {
        "id": lead["id"],
        "qualified": qualified,
        "score": score,
        "researchSummary": research_summary,
        "qualificationReasons": reasons,
    }


def qualify_batch(batch_id):
    """
    Qualify the requested batch using the latest uploaded batch.
    The current qualification pipeline only processes the most recent batch.
    """
    try:
        latest_batch = get_latest_batch(g.current_user["id"])
        if not latest_batch:
            return jsonify({"error": "No batches found"}), 404

        if str(latest_batch["_id"]) != batch_id:
            print(f"WARNING: Requested batch {batch_id}, but processing latest batch {latest_batch['_id']}")

        service = LeadQualificationService(g.current_user["id"])
        qualification_result = service.qualify_leads()

        if not qualification_result.get("success"):
            return jsonify({
                "error": qualification_result.get("error", "Qualification failed"),
                "stats": qualification_result.get("stats", {})
            }), 400

        # Convert qualification results to match leadModel format and save to leads
        qualified_leads_with_data = []
        for result in qualification_result.get("results", []):
            lead_data = {
                "id": result.get("lead_id"),
                "score": result.get("score"),
                "status": "strong_qualified" if result.get("status") == "strong_qualified" else ("moderate" if result.get("status") == "moderate" else "rejected"),
                "qualified": result.get("status") in ["strong_qualified", "moderate"],
                "researchSummary": result.get("qualification_reason", ""),
                "qualificationReasons": [result.get("qualification_reason", "")],
                "qualificationReason": result.get("qualification_reason", ""),
                "strengths": result.get("strengths", []),
                "weaknesses": result.get("weaknesses", []),
                "recommendedAction": result.get("recommended_action", ""),
                "priorityLevel": result.get("priority_level", "medium"),
                "aiGeneratedReasoning": result.get("ai_reasoning", ""),
            }
            qualified_leads_with_data.append(lead_data)

        # Save qualification data to leads collection
        mark_leads_qualified(g.current_user["id"], str(latest_batch["_id"]), qualified_leads_with_data)

        results = get_qualification_results(g.current_user["id"], str(latest_batch["_id"]))
        qualified_leads = list_leads(g.current_user["id"], batch_id=str(latest_batch["_id"]))

        update_batch_status(
            g.current_user["id"],
            str(latest_batch["_id"]),
            "qualified",
            qualification_result["stats"]
        )

        return jsonify({
            "success": True,
            "batch": sanitize_batch_doc(latest_batch),
            "leads": [sanitize_lead_doc(lead) for lead in qualified_leads],
            "qualification_results": [sanitize_qualification_result(r) for r in results],
            "stats": qualification_result["stats"],
            "message": f"Successfully qualified {len(results)} leads from latest batch"
        }), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Qualification failed: {str(e)}"}), 500


def get_qualification_results_for_batch(batch_id):
    """
    Fetch qualification results for a specific batch.
    Returns AI-generated scores, reasoning, and recommendations.
    """
    try:
        batch = get_batch(g.current_user["id"], batch_id)
        if not batch:
            return jsonify({"error": "Batch not found"}), 404

        results = get_qualification_results(g.current_user["id"], batch_id)

        return jsonify({
            "success": True,
            "batch_id": batch_id,
            "batch_name": batch.get("fileName"),
            "results": [sanitize_qualification_result(r) for r in results],
            "total_results": len(results),
        }), 200

    except Exception:
        traceback.print_exc()
        return jsonify({"error": "Failed to fetch qualification results"}), 500
